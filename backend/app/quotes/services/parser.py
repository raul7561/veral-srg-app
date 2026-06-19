import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.quotes.models import ParseResponse, QuoteLine


REQUIRED_COLUMNS = {
    "quantity": "Quantity",
    "part_name": "Part Name",
    "description": "Description",
}

OPTIONAL_COLUMNS = {
    "weight_each": "Weight Each",
    "line_item_note": "Line Item Note",
    "availability": "Availability",
    "unit_price_usd": "Unit Price in USD",
}

NO_STOCK_VALUES = {
    "Contact Dealer",
    "Contactar dealer",
    "Contactar al Dealer",
}


def parse_export(file_path: str) -> ParseResponse:
    suffix = Path(file_path).suffix.lower()

    if suffix == ".csv":
        rows = _read_csv(file_path)
    elif suffix in {".xlsx", ".xltm", ".xlsm"}:
        rows = _read_workbook(file_path)
    else:
        raise ValueError("Unsupported export format. Expected .csv, .xlsx, .xltm, or .xlsm.")

    return _transform_rows(rows)


def find_column(headers: list[str], nombre: str) -> str | None:
    target = nombre.lower()
    for header in headers:
        if header.lower() == target:
            return header
    for header in headers:
        if target in header.lower():
            return header
    return None


def clean_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)

    cleaned = str(value).strip().strip('"').strip("'").replace(",", "")
    if cleaned == "":
        return None

    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_weight(value: Any) -> float | None:
    if value is None:
        return None

    first_token = str(value).strip().split(" ", 1)[0]
    return clean_number(first_token)


def extract_replacement(note: str) -> str | None:
    match = re.search(r"Replaces Part\s*#\s*\(([^)]+)\)", note)
    if match:
        return match.group(1).strip()
    return None


def _read_csv(file_path: str) -> list[dict[str, Any]]:
    with open(file_path, newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.DictReader(csv_file)
        return [dict(row) for row in reader]


def _read_workbook(file_path: str) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return []

    headers = [_string_value(header) for header in header_row]
    parsed_rows = []
    for row in rows:
        parsed_rows.append(
            {
                header: value
                for header, value in zip(headers, row, strict=False)
            }
        )

    workbook.close()
    return parsed_rows


def _transform_rows(rows: list[dict[str, Any]]) -> ParseResponse:
    if not rows:
        return ParseResponse(lines=[], lines_generated=0, lines_discarded=0)

    headers = list(rows[0].keys())
    columns = _map_columns(headers)

    lines = []
    lines_discarded = 0

    for row in rows:
        quantity_value = clean_number(row.get(columns["quantity"]))
        if quantity_value is None or quantity_value <= 0:
            lines_discarded += 1
            continue

        part_name = _string_value(row.get(columns["part_name"])).strip()
        if part_name == "":
            lines_discarded += 1
            continue

        # Deteccion de fila corrida: una coma sin escapar en Description empuja
        # el resto de columnas una posicion. Señal: las columnas que en un export
        # sano vienen siempre vacias traen texto.
        shifted = False
        for guard_name in ("Customer Part Number", "Customer Item Number", "AssetID/Serial number"):
            guard_header = find_column(headers, guard_name)
            if guard_header and _string_value(row.get(guard_header)).strip():
                shifted = True
                break

        availability_raw = _string_value(_get_optional(row, columns, "availability")).strip()
        no_stock = availability_raw in NO_STOCK_VALUES

        raw_note = _string_value(_get_optional(row, columns, "line_item_note")).strip()
        replaces_part_number = extract_replacement(raw_note)
        if shifted:
            notes = (
                "Descripción corrida a otra celda, esto afectó el costo y la "
                "disponibilidad. Verificar quote exportada de Madisa."
            )
            replaces_part_number = None
        elif no_stock:
            notes = "No stock"
        elif availability_raw == "In Stock - Built To Order":
            notes = "Built to Order"
        elif replaces_part_number is not None:
            notes = ""
        else:
            notes = raw_note

        is_quotable = not (no_stock or shifted)

        lines.append(
            QuoteLine(
                item_number=len(lines) + 1,
                brand="CAT",
                quantity=int(quantity_value),
                part_number=part_name,
                description=_string_value(row.get(columns["description"])).strip(),
                madisa_cost=clean_number(_get_optional(row, columns, "unit_price_usd")) or 0.0,
                unit_price=None,
                minimum_qty=None,
                replaces_part_number=replaces_part_number,
                is_quotable=is_quotable,
                unit_weight=None if no_stock else parse_weight(_get_optional(row, columns, "weight_each")),
                notes=notes,
                availability_raw=availability_raw,
            )
        )

    return ParseResponse(
        lines=lines,
        lines_generated=len(lines),
        lines_discarded=lines_discarded,
    )


def _map_columns(headers: list[str]) -> dict[str, str | None]:
    columns: dict[str, str | None] = {}

    for key, name in REQUIRED_COLUMNS.items():
        header = find_column(headers, name)
        if header is None:
            raise ValueError(f"Missing required column: {name}")
        columns[key] = header

    for key, name in OPTIONAL_COLUMNS.items():
        columns[key] = find_column(headers, name)

    return columns


def _get_optional(row: dict[str, Any], columns: dict[str, str | None], key: str) -> Any:
    header = columns.get(key)
    if header is None:
        return None
    return row.get(header)


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)
