"""
Genera el Excel del quote a partir del template Client_quote.

Reglas:
- Parte del template (conserva encabezado, colores, formato de las notas).
- Escribe VALORES, nunca formulas (el calculo ya lo hizo el sistema).
- La tabla se ajusta al numero de lineas: recorta si hay menos, expande
  clonando formato si hay mas.
- El logo se inyecta como imagen (el template lo guarda como rich data, que
  openpyxl no preserva). Mismo PNG que el PDF.
- Las notas legales se despliegan una linea por fila, para que se lean.
- NUNCA incluye el costo MADISA (documento del cliente, igual que el PDF).
- Linea sin stock (unit_price None): UNIT PRICE y TOTAL PRICE van vacias.
"""
import copy
import io
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

_ASSETS = Path(__file__).parent / "assets"
_TEMPLATE = _ASSETS / "quote_template.xlsx"
_LOGO = _ASSETS / "srg_logo.png"

_FIRST_LINE_ROW = 11
_LAST_TEMPLATE_LINE_ROW = 56
_TEMPLATE_CAPACITY = _LAST_TEMPLATE_LINE_ROW - _FIRST_LINE_ROW + 1  # 46


def _fmt_date(raw) -> str:
    if not raw:
        return ""
    s = str(raw)
    parts = s.split("-")
    if len(parts) == 3:
        y, m, d = parts
        return f"{m}-{d}-{y}"
    return s


def _copy_row_style(ws, src_row: int, dst_row: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, 12):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        if s.has_style:
            d.font = copy.copy(s.font)
            d.border = copy.copy(s.border)
            d.fill = copy.copy(s.fill)
            d.number_format = s.number_format
            d.protection = copy.copy(s.protection)
            d.alignment = copy.copy(s.alignment)


def generate_quote_excel(quote: dict) -> bytes:
    lines = quote.get("lines", [])
    n = len(lines)

    wb = load_workbook(_TEMPLATE)
    ws = wb["Client_quote"]

    legal_text = ws.cell(row=59, column=1).value or ""

    ws["A2"] = None
    ws["J2"] = quote.get("quote_number", "")
    ws["J3"] = _fmt_date(quote.get("quote_date"))
    ws["J4"] = quote.get("sales_rep_name", "")
    ws["J5"] = quote.get("client_name", "")

    if n > _TEMPLATE_CAPACITY:
        extra = n - _TEMPLATE_CAPACITY
        insert_at = _LAST_TEMPLATE_LINE_ROW
        ws.insert_rows(insert_at, amount=extra)
        for i in range(extra):
            _copy_row_style(ws, _FIRST_LINE_ROW, insert_at + i)
    elif n < _TEMPLATE_CAPACITY:
        first_empty = _FIRST_LINE_ROW + n
        ws.delete_rows(first_empty, amount=_TEMPLATE_CAPACITY - n)

    total_row = _FIRST_LINE_ROW + n
    shipping_row = total_row + 1
    legal_row = total_row + 2

    for r in range(_FIRST_LINE_ROW, _FIRST_LINE_ROW + n):
        ws.cell(row=r, column=7).value = None
        ws.cell(row=r, column=9).value = None

    for idx, ln in enumerate(lines):
        r = _FIRST_LINE_ROW + idx
        up = ln.get("unit_price")
        qty = ln.get("quantity") or 0
        uw = ln.get("unit_weight")
        sin_stock = up is None
        ws.cell(row=r, column=1, value=ln.get("item_number"))
        ws.cell(row=r, column=2, value=ln.get("brand"))
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value=ln.get("part_number"))
        ws.cell(row=r, column=5, value=ln.get("description"))
        ws.cell(row=r, column=6, value=None if sin_stock else up)
        ws.cell(row=r, column=7, value=None if sin_stock else round(up * qty, 2))
        ws.cell(row=r, column=8, value=uw)
        ws.cell(row=r, column=9, value=None if uw is None else round(uw * qty, 2))
        ws.cell(row=r, column=10, value=ln.get("minimum_qty"))
        ws.cell(row=r, column=11, value=ln.get("notes") or None)

    ws.cell(row=total_row, column=7, value=round(quote.get("total_amount") or 0, 2))
    total_weight = sum((ln.get("unit_weight") or 0) * (ln.get("quantity") or 0) for ln in lines)
    ws.cell(row=total_row, column=9, value=round(total_weight, 2))

    ship = quote.get("shipping_cost")
    ws.cell(row=shipping_row, column=7, value=ship if ship is not None else "Pendiente")

    # Notas legales: una linea por fila
    for m in list(ws.merged_cells.ranges):
        if m.min_row == legal_row and m.max_row == legal_row:
            ws.unmerge_cells(str(m))
    ws.cell(row=legal_row, column=1).value = None
    ws.row_dimensions[legal_row].height = 15.8

    renglones = legal_text.split("\n")
    if len(renglones) > 1:
        ws.insert_rows(legal_row + 1, amount=len(renglones) - 1)
    for i, texto in enumerate(renglones):
        r = legal_row + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(row=r, column=1, value=texto)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.row_dimensions[r].height = 15.8

    if _LOGO.exists():
        img = XLImage(str(_LOGO))
        img.anchor = "A2"
        ws.add_image(img)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
